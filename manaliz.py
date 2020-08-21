import os
import text_analiz
import clean_text
import openpyxl
import winsound

def get_sheet_data(sheet_name):
    '''
    Получить данные с листа excel
    :param sheet_data: лист excel с которого нужно прочитать данные в первом столбце
    :return: данные
    '''
    str="is a text"
    row=2
    data=[]
    while len(str)>0:
        str = sheet_name.cell(row, 1).value
        if str=="None":
            break
        if str is None:
            break
        #print(str)
        data.append(str)
        row+=1
    return  data

def split_word(data_mas):
    '''
    Выделение слов - фраз пок которым анализируем
    :param data_mas: даннын для нализа
    :return: массив массивов слов
    '''
    data=[]
    for item in data_mas:
        split_words = text_analiz.getwords(item)
        data.append(split_words)
    return data

def clear_sheet( sheet_name, x0,y0, x, y):
    '''

    :param sheet_name:
    :param x0:
    :param y0:
    :param x:
    :param y:
    :return:
    '''
    for xx in range(x0,x+1):
        for yy in range(y0, y + 1):
            sheet_name.cell(xx, yy).value=""

def analiz(UPLOAD_FOLDER , filename):
    # ofn = "template.xlsx"
    # path_template = os.path.abspath(u'./' + ofn)
    path_template = UPLOAD_FOLDER + '\\' + filename

    # читаем excel-файл
    wb = openpyxl.load_workbook(path_template)

    # печатаем список листов
    # sheets = wb.sheetnames
    # for sheet in sheets:
    #     print(sheet)

    #Очищаем необходимые листы
    # clear_sheet(wb['разбиение по частям'],2,2,1000,40)
    # clear_sheet(wb['разбиение по частям (отфильтр)'],2,2,1000,40)
    # clear_sheet(wb['разбиение по частям (лемматиз)'],2,2,1000,40)
    # clear_sheet(wb['кластеры'],1,3,1000,40)


    sheet_key_words = wb['ключевые слова']
    key_words=get_sheet_data(sheet_key_words)

    sheet_clusters = wb['кластеры']
    for x in range(4,len(key_words)+4):
        sheet_clusters.cell(1, x).value = key_words[x-4]

    sheet_data_words = wb['данные для анализа']
    data_words=get_sheet_data(sheet_data_words)

    #Выделение слов - фраз пок которым анализируем
    data_words = split_word(data_words)
    # print(data_words)
    sheet_razb=wb['разбиение по частям']
    row=2
    for item1 in data_words:
        cow=2
        if type(item1) in (tuple, list):
            for item2 in item1:
                sheet_razb.cell(row, cow).value = item2
                cow+=1
        else:
            sheet_razb.cell(row, cow).value = item1
        row+=1

    #Фильтрация ненужных символов
    sheet_razb_otfiltr=wb['разбиение по частям (отфильтр)']
    row=2
    for item1 in data_words:
        cow=2
        if type(item1) in (tuple, list):
            for item2 in item1:
                data_words[row-2][cow-2] = clean_text.delete_bad_string(item2)
                sheet_razb_otfiltr.cell(row, cow).value = data_words[row-2][cow-2]
                cow+=1
        else:
            data_words[row - 2] = clean_text.delete_bad_string(item1)
            sheet_razb_otfiltr.cell(row, cow).value = data_words[row - 2]
        row+=1
    #print(data_words)

    lkey_words = text_analiz.d1lemmatize(key_words)
    #print(lkey_words)

    # Лематизация разделенных фраз
    ldata_words = text_analiz.d2lemmatize(data_words)
    #print(ldata_words)
    sheet_razb_lem=wb['разбиение по частям (лемматиз)']
    row=2
    for item1 in ldata_words:
        cow=2
        if type(item1) in (tuple, list):
            for item2 in item1:
                for item3 in item2:
                    sheet_razb_lem.cell(row, cow).value = item3
                    cow+=1
        else:
            sheet_razb_lem.cell(row, cow).value = item1
        row+=1

    # кластеры
    # сравнение по словам есть или нет слово
    row_num=2
    for item1 in ldata_words:
        for item2 in item1:
            cow_num=4
            for item3 in lkey_words:
                # print(item3)
                set1=item3
                set2=item2
                set1 = set(set1)
                set2 = set(set2)
                cmp = list(set1 & set2)
                if len(cmp)>0:
                    sheet_clusters.cell(row_num, cow_num).value = 1
                cow_num += 1
        row_num+=1
    wb.save(path_template)
    winsound.Beep(2500, 300)



if __name__ == "__main__":
    ofn = "template.xlsx"
    path_template = os.path.abspath(u'./' + ofn)
    analiz(path_template,"")


